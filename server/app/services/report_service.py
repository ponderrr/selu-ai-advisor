from io import BytesIO
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.models.user import User
from app.services.progress_service import calculate_academic_analytics, calculate_performance_metrics

def generate_pdf_report(user: User, analytics_data: dict) -> BytesIO:
    """Generate a PDF academic report."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30
    )
    elements.append(Paragraph(f"Academic Report for {user.first_name} {user.last_name}", title_style))
    elements.append(Paragraph(f"W-Number: {user.w_number}", styles['Normal']))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Overall Performance
    elements.append(Paragraph("Overall Performance", styles['Heading2']))
    overall = analytics_data['overall_metrics']
    overall_data = [
        ["Metric", "Value"],
        ["Average GPA", f"{overall['average_gpa']:.2f}"],
        ["Highest GPA", f"{overall['highest_gpa']:.2f}"],
        ["Lowest GPA", f"{overall['lowest_gpa']:.2f}"],
        ["Total Credits Earned", f"{overall['total_credits_earned']:.1f}"],
        ["Improvement Rate", f"{overall['improvement_rate']:.1f}%" if overall['improvement_rate'] else "N/A"]
    ]
    overall_table = Table(overall_data, colWidths=[3*inch, 2*inch])
    overall_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(overall_table)
    elements.append(Spacer(1, 20))

    # Grade Distribution
    elements.append(Paragraph("Grade Distribution", styles['Heading2']))
    grade_data = [["Grade", "Count", "Percentage"]] + [
        [item['grade'], str(item['count']), f"{item['percentage']:.1f}%"]
        for item in overall['grade_distribution']
    ]
    grade_table = Table(grade_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
    grade_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(grade_table)
    elements.append(Spacer(1, 20))

    # GPA Trend
    elements.append(Paragraph("GPA Trend", styles['Heading2']))
    trend_data = [["Semester", "Year", "Semester GPA", "Cumulative GPA", "Credits Earned"]] + [
        [
            item['semester'],
            str(item['year']),
            f"{item['semester_gpa']:.2f}",
            f"{item['cumulative_gpa']:.2f}",
            f"{item['credits_earned']:.1f}"
        ]
        for item in overall['gpa_trend']
    ]
    trend_table = Table(trend_data, colWidths=[1.2*inch, 1*inch, 1.2*inch, 1.2*inch, 1.2*inch])
    trend_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(trend_table)

    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer

def generate_excel_report(user: User, analytics_data: dict) -> BytesIO:
    """Generate an Excel academic report."""
    buffer = BytesIO()
    wb = Workbook()
    
    # Overall Performance Sheet
    ws_overall = wb.active
    ws_overall.title = "Overall Performance"
    
    # Title
    ws_overall['A1'] = f"Academic Report for {user.first_name} {user.last_name}"
    ws_overall['A2'] = f"W-Number: {user.w_number}"
    ws_overall['A3'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # Overall Metrics
    overall = analytics_data['overall_metrics']
    ws_overall['A5'] = "Overall Performance"
    ws_overall['A6'] = "Metric"
    ws_overall['B6'] = "Value"
    
    metrics = [
        ("Average GPA", f"{overall['average_gpa']:.2f}"),
        ("Highest GPA", f"{overall['highest_gpa']:.2f}"),
        ("Lowest GPA", f"{overall['lowest_gpa']:.2f}"),
        ("Total Credits Earned", f"{overall['total_credits_earned']:.1f}"),
        ("Improvement Rate", f"{overall['improvement_rate']:.1f}%" if overall['improvement_rate'] else "N/A")
    ]
    
    for i, (metric, value) in enumerate(metrics, start=7):
        ws_overall[f'A{i}'] = metric
        ws_overall[f'B{i}'] = value
    
    # Grade Distribution Sheet
    ws_grades = wb.create_sheet("Grade Distribution")
    ws_grades['A1'] = "Grade Distribution"
    ws_grades['A2'] = "Grade"
    ws_grades['B2'] = "Count"
    ws_grades['C2'] = "Percentage"
    
    for i, grade_data in enumerate(overall['grade_distribution'], start=3):
        ws_grades[f'A{i}'] = grade_data['grade']
        ws_grades[f'B{i}'] = grade_data['count']
        ws_grades[f'C{i}'] = f"{grade_data['percentage']:.1f}%"
    
    # GPA Trend Sheet
    ws_trend = wb.create_sheet("GPA Trend")
    ws_trend['A1'] = "GPA Trend"
    headers = ["Semester", "Year", "Semester GPA", "Cumulative GPA", "Credits Earned"]
    for col, header in enumerate(headers, start=1):
        ws_trend[f'{get_column_letter(col)}2'] = header
    
    for i, trend in enumerate(overall['gpa_trend'], start=3):
        ws_trend[f'A{i}'] = trend['semester']
        ws_trend[f'B{i}'] = trend['year']
        ws_trend[f'C{i}'] = f"{trend['semester_gpa']:.2f}"
        ws_trend[f'D{i}'] = f"{trend['cumulative_gpa']:.2f}"
        ws_trend[f'E{i}'] = f"{trend['credits_earned']:.1f}"
    
    # Style the sheets
    for ws in [ws_overall, ws_grades, ws_trend]:
        # Title styling
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'].font = Font(size=12)
        ws['A3'].font = Font(size=12)
        
        # Header styling
        for row in ws.iter_rows(min_row=2, max_row=2):
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
        
        # Data styling
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generate_report(user: User, db, format: str = 'pdf') -> tuple[BytesIO, str, str]:
    """Generate an academic report in the specified format."""
    analytics_data = calculate_academic_analytics(user, db)
    
    if format.lower() == 'excel':
        buffer = generate_excel_report(user, analytics_data)
        mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        file_extension = 'xlsx'
    else:  # default to PDF
        buffer = generate_pdf_report(user, analytics_data)
        mime_type = 'application/pdf'
        file_extension = 'pdf'
    
    return buffer, mime_type, file_extension 